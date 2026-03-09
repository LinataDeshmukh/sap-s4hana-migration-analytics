# =============================================================================
# profiling/export_bronze_silver.py
#
# PURPOSE:
#   Exports bronze and silver tables to Excel for side-by-side comparison.
#   Each pair gets one Excel file with 3 sheets:
#     - Bronze_Raw     → raw messy data sorted by load_id
#     - Silver_Cleaned → cleaned data sorted by load_id, color coded
#     - DQ_Summary     → issue counts per flag
#
# HOW TO RUN:
#   python profiling/export_bronze_silver.py
#
# OUTPUT:
#   profiling/comparisons/comparison_<table>.xlsx
# =============================================================================

import sys
import os
import pandas as pd
from datetime import datetime
from sqlalchemy import create_engine
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config.settings import DB_CONFIG

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "comparisons"
)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# =============================================================================
# TABLE PAIRS TO EXPORT
# Add new pair here every time a new silver model is built
# Format: (bronze_schema, bronze_table, silver_schema, silver_table, label)
# =============================================================================

EXPORT_PAIRS = [
    (
        "precision_mfg_bronze",
        "bronze_material_master",
        "precision_mfg_bronze_silver",
        "silver_material_master",
        "material_master"
    ),
    (
        "precision_mfg_bronze",
        "bronze_vendor_master",
        "precision_mfg_bronze_silver",
        "silver_vendor_master",
        "vendor_master"
    ),
    (
        "precision_mfg_bronze",
        "bronze_mrp_parameters",
        "precision_mfg_bronze_silver",
        "silver_mrp_parameters",
        "mrp_parameters"
    ),
    (
        "precision_mfg_bronze",
        "bronze_material_plant_data",
        "precision_mfg_bronze_silver",
        "silver_material_plant_data",
        "material_plant_data"
    ),
    (
        "precision_mfg_bronze",
        "bronze_equipment_master",
        "precision_mfg_bronze_silver",
        "silver_equipment_master",
        "equipment_master"
    ),
    (
        "precision_mfg_bronze",
        "bronze_cost_centers",
        "precision_mfg_bronze_silver",
        "silver_cost_centers",
        "cost_centers"
    ),
]

# =============================================================================
# COLUMNS TO HIGHLIGHT IN SILVER SHEET
# Blue background = cleaned value column
# =============================================================================

HIGHLIGHT_COLS = [
    # Material master
    "material_id_clean",
    "base_uom_clean",
    "material_type_clean",
    "material_group_clean",
    "created_date_clean",
    "changed_date_clean",
    "gross_weight_num",
    "net_weight_num",
    "procurement_type_clean",
    # Vendor master
    "vendor_id_clean",
    "country_clean",
    "currency_clean",
    "vendor_status_clean",
    "vendor_type_clean",
    "payment_terms_clean",
    "language_clean",
    # Common
    "dq_score",
    # MRP parameters
    "mrp_type_clean",
    "lot_size_key_clean",
    "backward_scheduling_clean",
    "availability_check_clean",
    "planned_delivery_days_num",
    "safety_stock_num",
    "reorder_point_num",
    "goods_receipt_days_num",
    # Material plant data
    "mrp_type_clean",
    "price_control_clean",
    "lot_size_clean",
    "safety_stock_num",
    "standard_price_num",
    "moving_avg_price_num",
    "planned_delivery_days_num",
    "goods_receipt_days_num",
    "valid_from_clean",
    "created_date_clean",
    # Equipment master
    "active_flag_clean",
    "currency_clean",
    "manufacturer_clean",
    "acquisition_date_clean",
    "start_up_date_clean",
    "warranty_end_date_clean",
    "created_date_clean",
    "changed_date_clean",
    "acquisition_value_num",
    "year_constructed_num",
    # Cost centers
    "active_flag_clean",
    "currency_clean",
    "cost_center_type_clean",
    "department_clean",
    "valid_from_clean",
    "valid_to_clean",
    "created_date_clean",
    "changed_date_clean",
]

# =============================================================================
# FLAG COLUMNS
# Green = 0 (clean), Red = 1 (issue)
# =============================================================================

FLAG_COLS = [
    # Material master flags
    "flag_null_description",
    "flag_null_material_type",
    "flag_invalid_material_type",
    "flag_null_material_group",
    "flag_null_uom",
    "flag_nonstandard_uom",
    "flag_negative_gross_weight",
    "flag_negative_net_weight",
    "flag_weight_invalid",
    "flag_bad_created_date",
    "flag_changed_before_created",
    "flag_generic_user",
    "flag_nonstandard_material_id",
    "flag_duplicate",
    "flag_invalid_procurement_type",
    # Vendor master flags
    "flag_null_vendor_name",
    "flag_invalid_country",
    "flag_missing_bank_account",
    "flag_missing_bank_routing",
    "flag_missing_payment_terms",
    "flag_invalid_vendor_status",
    "flag_invalid_vendor_type",
    "flag_missing_currency",
    "flag_nonstandard_currency",
    "flag_missing_payment_method",
    "flag_nonstandard_country",
    "flag_duplicate_vendor",
    # MRP parameters flags
    "flag_invalid_plant",
    "flag_null_mrp_type",
    "flag_invalid_mrp_type",
    "flag_null_lot_size_key",
    "flag_null_safety_stock",
    "flag_invalid_lead_time",
    "flag_negative_safety_stock",
    "flag_negative_reorder_point",
    "flag_invalid_consumption_mode",
    # Material plant data flags
    "flag_invalid_plant",
    "flag_null_mrp_type",
    "flag_invalid_mrp_type",
    "flag_invalid_price_control",
    "flag_null_valuation_class",
    "flag_null_safety_stock",
    "flag_negative_standard_price",
    "flag_negative_moving_avg_price",
    "flag_invalid_lot_size",
    "flag_invalid_lead_time",
    "flag_bad_valid_from",
    "flag_bad_created_date",
    # Equipment master flags
    "flag_invalid_plant",
    "flag_invalid_company_code",
    "flag_invalid_active_flag",
    "flag_invalid_cost_center",
    "flag_null_cost_center",
    "flag_null_description",
    "flag_future_construction",
    "flag_old_construction_year",
    "flag_warranty_before_startup",
    "flag_changed_before_created",
    "flag_generic_user",
    "flag_nonstandard_currency",
    "flag_duplicate_serial",
    # Cost center flags
    "flag_invalid_company_code",
    "flag_invalid_controlling_area",
    "flag_invalid_plant",
    "flag_invalid_active_flag",
    "flag_invalid_cc_type",
    "flag_null_responsible_person",
    "flag_null_profit_center",
    "flag_invalid_validity_dates",
    "flag_expired_but_active",
    "flag_bad_created_date",
    "flag_changed_before_created",
    "flag_generic_user",
]

# =============================================================================
# STYLING HELPERS
# =============================================================================

def style_header(ws, num_cols):
    """Dark blue header with white bold text."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"


def style_silver_sheet(ws, df):
    """
    Colors the silver sheet:
    - Blue  = cleaned value columns
    - Green = flag = 0 (clean)
    - Red   = flag = 1 (issue)
    - Green/Amber/Red = dq_score
    """
    col_names = list(df.columns)

    for col_idx, col_name in enumerate(col_names, start=1):
        col_letter = get_column_letter(col_idx)

        # Blue background for cleaned value columns
        if col_name in HIGHLIGHT_COLS and col_name != "dq_score":
            for row in range(2, len(df) + 2):
                ws.cell(row=row, column=col_idx).fill = PatternFill(
                    start_color="DDEEFF", end_color="DDEEFF", fill_type="solid"
                )

        # Green/Red for flag columns
        if col_name in FLAG_COLS:
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                if str(cell.value) == "1":
                    cell.fill = PatternFill(
                        start_color="FFC7CE",
                        end_color="FFC7CE",
                        fill_type="solid"
                    )
                elif str(cell.value) == "0":
                    cell.fill = PatternFill(
                        start_color="C6EFCE",
                        end_color="C6EFCE",
                        fill_type="solid"
                    )

        # Green/Amber/Red for dq_score
        if col_name == "dq_score":
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    score = float(cell.value)
                    if score >= 95:
                        color = "C6EFCE"
                    elif score >= 80:
                        color = "FFEB9C"
                    else:
                        color = "FFC7CE"
                    cell.fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )
                except (TypeError, ValueError):
                    pass

        ws.column_dimensions[col_letter].width = 20

    ws.freeze_panes = "A2"


# =============================================================================
# EXPORT ONE PAIR
# =============================================================================

def export_pair(engine, bronze_schema, bronze_table,
                silver_schema, silver_table, label):
    """
    Exports one bronze/silver table pair to Excel.
    Sorted by load_id so bronze row 1 = silver row 1.
    """
    output_file = os.path.join(OUTPUT_DIR, f"comparison_{label}.xlsx")
    print(f"\nExporting: {label}")

    # ── Read bronze sorted by load_id ────────────────────────────────────────
    try:
        df_bronze = pd.read_sql(
            f"""
            SELECT *
            FROM `{bronze_schema}`.`{bronze_table}`
            ORDER BY load_id
            LIMIT 500
            """,
            engine
        )

        # Move _source_file and _batch_id to first two columns
        priority_cols = [
            c for c in ["_source_file", "_batch_id"]
            if c in df_bronze.columns
        ]
        other_cols = [
            c for c in df_bronze.columns
            if c not in priority_cols
        ]
        df_bronze = df_bronze[priority_cols + other_cols]

        print(f"  Bronze: {len(df_bronze):,} rows, "
              f"{len(df_bronze.columns)} columns")

    except Exception as e:
        print(f"  Bronze FAILED: {e}")
        df_bronze = pd.DataFrame()

    # ── Read silver sorted by load_id ─────────────────────────────────────────
    try:
        df_silver = pd.read_sql(
            f"""
            SELECT *
            FROM `{silver_schema}`.`{silver_table}`
            ORDER BY load_id
            LIMIT 500
            """,
            engine
        )
        print(f"  Silver: {len(df_silver):,} rows, "
              f"{len(df_silver.columns)} columns")

    except Exception as e:
        print(f"  Silver FAILED (model not built yet?): {e}")
        df_silver = pd.DataFrame()

    # ── Write Excel ───────────────────────────────────────────────────────────
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        # ── Sheet 1: Bronze Raw ───────────────────────────────────────────────
        if not df_bronze.empty:
            df_bronze.to_excel(
                writer, sheet_name="Bronze_Raw", index=False
            )
            ws = writer.sheets["Bronze_Raw"]
            style_header(ws, len(df_bronze.columns))
            for col_idx in range(1, len(df_bronze.columns) + 1):
                ws.column_dimensions[
                    get_column_letter(col_idx)
                ].width = 20
            ws.freeze_panes = "A2"

        # ── Sheet 2: Silver Cleaned ───────────────────────────────────────────
        if not df_silver.empty:
            df_silver.to_excel(
                writer, sheet_name="Silver_Cleaned", index=False
            )
            ws = writer.sheets["Silver_Cleaned"]
            style_header(ws, len(df_silver.columns))
            style_silver_sheet(ws, df_silver)

        # ── Sheet 3: DQ Summary ───────────────────────────────────────────────
        if not df_silver.empty:
            summary_rows = []

            for flag in FLAG_COLS:
                if flag in df_silver.columns:
                    issue_count = int(
                        df_silver[flag].astype(str).eq("1").sum()
                    )
                    clean_count = int(
                        df_silver[flag].astype(str).eq("0").sum()
                    )
                    summary_rows.append({
                        "Flag":        flag,
                        "Issue Count": issue_count,
                        "Clean Count": clean_count,
                        "Issue %": round(
                            issue_count / len(df_silver) * 100, 1
                        ) if len(df_silver) > 0 else 0,
                        "Severity": (
                            "CRITICAL" if issue_count / len(df_silver) > 0.5
                            else "HIGH"    if issue_count / len(df_silver) > 0.2
                            else "MEDIUM"  if issue_count / len(df_silver) > 0.05
                            else "LOW"
                        ) if len(df_silver) > 0 else "N/A"
                    })

            if summary_rows:
                df_summary = pd.DataFrame(summary_rows)

                if "dq_score" in df_silver.columns:
                    avg_score = round(
                        pd.to_numeric(
                            df_silver["dq_score"], errors="coerce"
                        ).mean(), 1
                    )
                    df_summary.loc[len(df_summary)] = {
                        "Flag":        "── AVG DQ SCORE ──",
                        "Issue Count": "",
                        "Clean Count": "",
                        "Issue %":     avg_score,
                        "Severity":    (
                            "READY"     if avg_score >= 95
                            else "CAUTION"  if avg_score >= 80
                            else "NOT READY"
                        )
                    }

                df_summary.to_excel(
                    writer, sheet_name="DQ_Summary", index=False
                )
                ws = writer.sheets["DQ_Summary"]
                style_header(ws, len(df_summary.columns))
                for col_idx in range(1, len(df_summary.columns) + 1):
                    ws.column_dimensions[
                        get_column_letter(col_idx)
                    ].width = 25

    print(f"  Saved: {output_file}")
    return output_file


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":

    started_at = datetime.now()

    print("=" * 65)
    print("SAP Migration - Bronze vs Silver Comparison Exporter")
    print(f"Started : {started_at.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Output  : {OUTPUT_DIR}")
    print("=" * 65)

    engine = create_engine(
        f"mysql+pymysql://{DB_CONFIG['user']}:{DB_CONFIG['password']}"
        f"@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}"
        f"?charset=utf8mb4"
    )

    exported = []
    for (bronze_schema, bronze_table,
         silver_schema, silver_table, label) in EXPORT_PAIRS:
        f = export_pair(
            engine,
            bronze_schema, bronze_table,
            silver_schema, silver_table,
            label
        )
        exported.append(f)

    duration = (datetime.now() - started_at).total_seconds()

    print("\n" + "=" * 65)
    print("EXPORT COMPLETE")
    print("=" * 65)
    for f in exported:
        print(f"  {f}")
    print(f"\nDuration: {duration:.1f}s")
    print("=" * 65)
    print("\nEach Excel file has 3 sheets:")
    print("  Bronze_Raw     → raw messy data (sorted by load_id)")
    print("  Silver_Cleaned → cleaned data (sorted by load_id)")
    print("                   Blue  = cleaned value columns")
    print("                   Green = flag 0 (clean)")
    print("                   Red   = flag 1 (issue)")
    print("  DQ_Summary     → issue counts and severity per flag")
    print("=" * 65)