# =============================================================================
# profiling/generate_dq_report.py
#
# PURPOSE:
#   Runs all data quality profiling queries against bronze_column_profile
#   and saves results to an Excel report with one sheet per analysis.
#
#   Output file: profiling/DQ_Assessment_Report.xlsx
#
#   Sheets:
#     1. Summary          → DQ summary by domain
#     2. High_Null_Cols   → Columns with highest null rates
#     3. Mixed_Types      → Columns with mixed data types
#     4. Table_Health     → Per table readiness status
#     5. All_Columns      → Full profile of every column
#     6. Top_Issues       → Top 20 issues to fix per domain
#
# HOW TO RUN:
#   python profiling/generate_dq_report.py
# =============================================================================

import sys
import os
import pandas as pd
from datetime import datetime
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from ingestion.db_connection import get_connection
from sqlalchemy import create_engine
from config.settings import DB_CONFIG

# ── Force UTF-8 on Windows console ───────────────────────────────────────────
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

OUTPUT_DIR  = os.path.join(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_FILE = os.path.join(
    OUTPUT_DIR,
    f"DQ_Assessment_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)

# =============================================================================
# QUERIES
# =============================================================================

QUERIES = {

    # ── Sheet 1: Domain Summary ───────────────────────────────────────────────
    "Domain_Summary": """
        SELECT
            domain                                          AS Domain,
            COUNT(DISTINCT table_name)                      AS Tables,
            MAX(total_rows)                                 AS Total_Rows,
            COUNT(*)                                        AS Columns_Profiled,
            SUM(has_nulls)                                  AS Cols_With_Nulls,
            SUM(high_null_rate)                             AS Cols_High_Null_Rate,
            SUM(has_blanks)                                 AS Cols_With_Blanks,
            SUM(mixed_types)                                AS Cols_Mixed_Types,
            SUM(is_constant)                                AS Constant_Cols,
            ROUND(AVG(populated_pct), 1)                    AS Avg_Populated_Pct,
            CASE
                WHEN AVG(populated_pct) >= 95 THEN 'READY'
                WHEN AVG(populated_pct) >= 80 THEN 'CAUTION'
                ELSE 'NOT READY'
            END                                             AS Readiness_Status
        FROM bronze_column_profile
        WHERE domain NOT IN ('Audit')
        GROUP BY domain
        ORDER BY Avg_Populated_Pct ASC
    """,

    # ── Sheet 2: Table Health ─────────────────────────────────────────────────
    "Table_Health": """
        SELECT
            table_name                                      AS Table_Name,
            domain                                          AS Domain,
            total_rows                                      AS Total_Rows,
            COUNT(*)                                        AS Total_Columns,
            SUM(has_nulls)                                  AS Cols_With_Nulls,
            SUM(high_null_rate)                             AS Cols_High_Null_Rate,
            SUM(has_blanks)                                 AS Cols_With_Blanks,
            SUM(mixed_types)                                AS Cols_Mixed_Types,
            SUM(all_null)                                   AS Completely_Null_Cols,
            ROUND(AVG(populated_pct), 1)                    AS Avg_Populated_Pct,
            ROUND(MIN(populated_pct), 1)                    AS Min_Populated_Pct,
            CASE
                WHEN AVG(populated_pct) >= 95 THEN 'READY'
                WHEN AVG(populated_pct) >= 80 THEN 'CAUTION'
                ELSE 'NOT READY'
            END                                             AS Readiness_Status
        FROM bronze_column_profile
        WHERE domain NOT IN ('Audit')
        GROUP BY table_name, domain, total_rows
        ORDER BY Avg_Populated_Pct ASC
    """,

    # ── Sheet 3: High Null Columns ────────────────────────────────────────────
    "High_Null_Columns": """
        SELECT
            domain                                          AS Domain,
            table_name                                      AS Table_Name,
            column_name                                     AS Column_Name,
            total_rows                                      AS Total_Rows,
            null_count                                      AS Null_Count,
            null_pct                                        AS Null_Pct,
            blank_count                                     AS Blank_Count,
            blank_pct                                       AS Blank_Pct,
            populated_count                                 AS Populated_Count,
            populated_pct                                   AS Populated_Pct,
            distinct_count                                  AS Distinct_Values,
            top_5_values                                    AS Top_5_Values,
            CASE
                WHEN null_pct = 100 THEN 'CRITICAL - All Null'
                WHEN null_pct >= 50  THEN 'CRITICAL - Mostly Null'
                WHEN null_pct >= 20  THEN 'HIGH - Many Nulls'
                ELSE 'MEDIUM - Some Nulls'
            END                                             AS Severity
        FROM bronze_column_profile
        WHERE null_pct > 5
          AND domain NOT IN ('Audit')
        ORDER BY null_pct DESC
    """,

    # ── Sheet 4: Mixed Type Columns ───────────────────────────────────────────
    "Mixed_Type_Columns": """
        SELECT
            domain                                          AS Domain,
            table_name                                      AS Table_Name,
            column_name                                     AS Column_Name,
            total_rows                                      AS Total_Rows,
            populated_count                                 AS Populated_Count,
            numeric_count                                   AS Numeric_Count,
            numeric_pct                                     AS Numeric_Pct,
            distinct_count                                  AS Distinct_Values,
            min_length                                      AS Min_Length,
            max_length                                      AS Max_Length,
            sample_values                                   AS Sample_Values,
            top_5_values                                    AS Top_5_Values
        FROM bronze_column_profile
        WHERE mixed_types = 1
          AND domain NOT IN ('Audit')
        ORDER BY domain, table_name, column_name
    """,

    # ── Sheet 5: Top Issues Per Domain ────────────────────────────────────────
    "Top_Issues": """
        SELECT
            domain                                          AS Domain,
            table_name                                      AS Table_Name,
            column_name                                     AS Column_Name,
            total_rows                                      AS Total_Rows,
            null_count                                      AS Null_Count,
            null_pct                                        AS Null_Pct,
            blank_count                                     AS Blank_Count,
            distinct_count                                  AS Distinct_Values,
            has_nulls                                       AS Has_Nulls,
            high_null_rate                                  AS High_Null_Rate,
            has_blanks                                      AS Has_Blanks,
            mixed_types                                     AS Mixed_Types,
            is_constant                                     AS Is_Constant,
            top_5_values                                    AS Top_5_Values,
            sample_values                                   AS Sample_Values,
            -- Issue description
            CASE
                WHEN null_pct >= 50  THEN 'CRITICAL: Column is mostly null'
                WHEN null_pct >= 20  THEN 'HIGH: High null rate - mandatory field?'
                WHEN mixed_types = 1 THEN 'MEDIUM: Mixed data types in column'
                WHEN is_constant = 1 THEN 'LOW: Only one distinct value'
                WHEN has_blanks = 1  THEN 'LOW: Contains blank strings'
                ELSE 'INFO: Some nulls present'
            END                                             AS Issue_Description,
            -- Recommendation
            CASE
                WHEN null_pct >= 50  THEN 'Collect missing data from business team'
                WHEN null_pct >= 20  THEN 'Review if field is mandatory for SAP load'
                WHEN mixed_types = 1 THEN 'Standardize data type before SAP upload'
                WHEN is_constant = 1 THEN 'Verify if column is needed'
                WHEN has_blanks = 1  THEN 'Replace blank strings with NULL'
                ELSE 'Monitor - low risk'
            END                                             AS Recommendation
        FROM bronze_column_profile
        WHERE (has_nulls = 1 OR mixed_types = 1 OR is_constant = 1)
          AND domain NOT IN ('Audit')
        ORDER BY
            CASE
                WHEN null_pct >= 50  THEN 1
                WHEN null_pct >= 20  THEN 2
                WHEN mixed_types = 1 THEN 3
                ELSE 4
            END,
            null_pct DESC
    """,

    # ── Sheet 6: Full Column Profile ──────────────────────────────────────────
    "Full_Column_Profile": """
        SELECT
            domain                                          AS Domain,
            table_name                                      AS Table_Name,
            column_name                                     AS Column_Name,
            column_position                                 AS Position,
            total_rows                                      AS Total_Rows,
            null_count                                      AS Null_Count,
            null_pct                                        AS Null_Pct,
            blank_count                                     AS Blank_Count,
            populated_count                                 AS Populated_Count,
            populated_pct                                   AS Populated_Pct,
            distinct_count                                  AS Distinct_Values,
            distinct_pct                                    AS Distinct_Pct,
            min_length                                      AS Min_Length,
            max_length                                      AS Max_Length,
            avg_length                                      AS Avg_Length,
            numeric_count                                   AS Numeric_Count,
            numeric_pct                                     AS Numeric_Pct,
            date_count                                      AS Date_Count,
            date_pct                                        AS Date_Pct,
            has_nulls                                       AS Has_Nulls,
            high_null_rate                                  AS High_Null_Rate,
            all_null                                        AS All_Null,
            is_constant                                     AS Is_Constant,
            high_cardinality                                AS High_Cardinality,
            has_blanks                                      AS Has_Blanks,
            mixed_types                                     AS Mixed_Types,
            top_5_values                                    AS Top_5_Values,
            sample_values                                   AS Sample_Values
        FROM bronze_column_profile
        ORDER BY domain, table_name, column_position
    """,
}


# =============================================================================
# EXCEL STYLING HELPERS
# =============================================================================

# Color scheme
COLORS = {
    "header_bg":    "1F4E79",   # dark blue header
    "header_font":  "FFFFFF",   # white text
    "ready":        "C6EFCE",   # green
    "caution":      "FFEB9C",   # amber
    "not_ready":    "FFC7CE",   # red
    "critical":     "FF0000",   # bright red
    "high":         "FF6600",   # orange
    "medium":       "FFCC00",   # yellow
    "low":          "99CC00",   # light green
    "alt_row":      "EBF3FB",   # light blue alternating row
}

def style_header_row(ws, row_num, num_cols):
    """Styles the header row with dark blue background and white bold text."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = PatternFill(
            start_color=COLORS["header_bg"],
            end_color=COLORS["header_bg"],
            fill_type="solid"
        )
        cell.font      = Font(bold=True, color=COLORS["header_font"], size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)


def style_status_cell(cell, value):
    """Colors a cell based on its readiness status value."""
    color_map = {
        "READY":     COLORS["ready"],
        "CAUTION":   COLORS["caution"],
        "NOT READY": COLORS["not_ready"],
        "CRITICAL - All Null":    COLORS["critical"],
        "CRITICAL - Mostly Null": COLORS["critical"],
        "HIGH - Many Nulls":      COLORS["high"],
        "MEDIUM - Some Nulls":    COLORS["medium"],
    }
    if value in color_map:
        cell.fill = PatternFill(
            start_color=color_map[value],
            end_color=color_map[value],
            fill_type="solid"
        )


def auto_fit_columns(ws, df):
    """Auto-fits column widths based on content."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        # Width = max of header length and longest value, capped at 50
        max_len = max(
            len(str(col_name)),
            df[col_name].astype(str).str.len().max() if len(df) > 0 else 0
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def write_sheet(writer, sheet_name, df, status_col=None):
    """
    Writes a DataFrame to an Excel sheet with formatting.

    Args:
        writer:     pd.ExcelWriter
        sheet_name: Name of the sheet
        df:         DataFrame to write
        status_col: Column name to apply status color coding (optional)
    """
    if df.empty:
        print(f"  {sheet_name}: No data")
        return

    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    # Style header
    style_header_row(ws, 1, len(df.columns))

    # Style data rows
    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(
                    start_color=COLORS["alt_row"],
                    end_color=COLORS["alt_row"],
                    fill_type="solid"
                )

            cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Apply status color if specified
        if status_col and status_col in df.columns:
            status_col_idx = df.columns.get_loc(status_col) + 1
            status_cell    = ws.cell(row=row_idx, column=status_col_idx)
            style_status_cell(status_cell, status_cell.value)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-fit columns
    auto_fit_columns(ws, df)

    print(f"  {sheet_name}: {len(df):,} rows written")


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":

    started_at = datetime.now()

    print("=" * 65)
    print("SAP Migration - Data Quality Assessment Report Generator")
    print(f"Started : {started_at.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Output  : {OUTPUT_FILE}")
    print("=" * 65)

    # ── Connect to MySQL ──────────────────────────────────────────────────────
    #conn = get_connection()
    engine = create_engine(
    f"mysql+pymysql://{DB_CONFIG['user']}:{DB_CONFIG['password']}"
    f"@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}"
    f"?charset=utf8mb4"
    )
    conn = engine.connect()

    # ── Run all queries and collect DataFrames ────────────────────────────────
    print("\nRunning profiling queries...")
    dataframes = {}

    for query_name, sql in QUERIES.items():
        try:
            df = pd.read_sql(sql, conn)
            dataframes[query_name] = df
            print(f"  {query_name}: {len(df):,} rows")
        except Exception as e:
            print(f"  {query_name}: FAILED - {e}")
            dataframes[query_name] = pd.DataFrame()

   
    conn.close()
    engine.dispose()

    # ── Write to Excel ────────────────────────────────────────────────────────
    print(f"\nWriting Excel report...")

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

        # Sheet 1: Domain Summary
        write_sheet(
            writer, "Domain_Summary",
            dataframes["Domain_Summary"],
            status_col="Readiness_Status"
        )

        # Sheet 2: Table Health
        write_sheet(
            writer, "Table_Health",
            dataframes["Table_Health"],
            status_col="Readiness_Status"
        )

        # Sheet 3: High Null Columns
        write_sheet(
            writer, "High_Null_Columns",
            dataframes["High_Null_Columns"],
            status_col="Severity"
        )

        # Sheet 4: Mixed Type Columns
        write_sheet(
            writer, "Mixed_Type_Columns",
            dataframes["Mixed_Type_Columns"]
        )

        # Sheet 5: Top Issues
        write_sheet(
            writer, "Top_Issues",
            dataframes["Top_Issues"]
        )

        # Sheet 6: Full Column Profile
        write_sheet(
            writer, "Full_Column_Profile",
            dataframes["Full_Column_Profile"]
        )

    # ── Summary ───────────────────────────────────────────────────────────────
    duration = (datetime.now() - started_at).total_seconds()

    print("\n" + "=" * 65)
    print("REPORT GENERATED SUCCESSFULLY")
    print("=" * 65)
    print(f"File    : {OUTPUT_FILE}")
    print(f"Sheets  : 6")
    print(f"Duration: {duration:.1f}s")
    print("=" * 65)
    print("\nOpen the Excel file to review:")
    print("  Sheet 1 - Domain_Summary    : Overall readiness by domain")
    print("  Sheet 2 - Table_Health      : Per table readiness status")
    print("  Sheet 3 - High_Null_Columns : Columns with most nulls")
    print("  Sheet 4 - Mixed_Type_Columns: Columns with type issues")
    print("  Sheet 5 - Top_Issues        : Prioritized fix list")
    print("  Sheet 6 - Full_Column_Profile: Every column profiled")
    print("=" * 65)